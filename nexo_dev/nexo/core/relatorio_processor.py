"""
Processador de Relatórios - Extrai dados de planilhas Excel e salva no banco de dados.
"""

import pandas as pd
import openpyxl
from datetime import datetime
from django.utils import timezone
from .models import (
    Relatorio,
    RelatorioGratificacoes,
    RelatorioOrgaosCentrais,
    RelatorioEfetivo,
)


def processar_relatorio(relatorio_obj):
    """
    Processa um arquivo de relatório Excel baseado no seu tipo.
    """
    try:
        arquivo_path = relatorio_obj.arquivo.path
        tipo_relatorio = relatorio_obj.tipo

        if tipo_relatorio == "gratificacoes":
            return processar_gratificacoes(relatorio_obj, arquivo_path)
        elif tipo_relatorio == "orgaos":
            return processar_orgaos_centrais(relatorio_obj, arquivo_path)
        elif tipo_relatorio in ["efetivo", "facilities"]:
            return processar_efetivo(relatorio_obj, arquivo_path)
        else:
            return (
                False,
                "Tipo de relatório não suportado para processamento automático.",
            )

    except Exception as e:
        return False, f"Erro ao processar relatório: {str(e)}"


def processar_gratificacoes(relatorio_obj, arquivo_path):
    """
    Processa planilha de gratificações e lotações.
    """
    try:
        # Ler o arquivo Excel
        df = pd.read_excel(arquivo_path)

        # Limpar dados existentes para este relatório
        RelatorioGratificacoes.objects.filter(relatorio=relatorio_obj).delete()

        registros_criados = 0

        # Mapear colunas da planilha para campos do modelo
        # Ordem das colunas conforme especificado pelo usuário:
        # Nome do Servidor, Matrícula SIAPE, CPF, Data de Nascimento, Idade, Sexo, Situação Funcional,
        # Cargo, Nível, Gsiste, Gsiste Nível, Função, Nível da Função, Atividade da Função,
        # Jornada de Trabalho, Unidade de Lotação, Secretaria da Lotação, UF, UORG de Exercício,
        # Unidade de Exercício, Coordenação, Diretoria, Secretaria, Órgão Origem, e-Mail Institucional,
        # Siape do Titular Chefe, CPF do Titular do Chefe, Siape do Substituto, CPF do Substituto

        for _, row in df.iterrows():
            try:
                # Pular linhas vazias ou inválidas
                if pd.isna(row.iloc[0]) or str(row.iloc[0]).strip() == "":
                    continue

                # Função auxiliar para obter valor da coluna de forma segura
                def get_value(index, default=""):
                    if len(row) > index and not pd.isna(row.iloc[index]):
                        return str(row.iloc[index]).strip()
                    return default

                def get_int_value(index, default=None):
                    if len(row) > index and not pd.isna(row.iloc[index]):
                        try:
                            return int(float(row.iloc[index]))
                        except:
                            return default
                    return default

                # Criar registro com mapeamento completo
                dados = {
                    "relatorio": relatorio_obj,
                    "nome_servidor": get_value(0),
                    "matricula_siape": get_value(1),
                    "cpf": get_value(2),
                    "idade": get_int_value(4),
                    "sexo": get_value(5),
                    "situacao_funcional": get_value(6),
                    "cargo": get_value(7),
                    "nivel": get_value(8),
                    "gsiste": get_value(9),
                    "gsiste_nivel": get_value(10),
                    "funcao": get_value(11),
                    "nivel_funcao": get_value(12),
                    "atividade_funcao": get_value(13),
                    "jornada_trabalho": get_value(14),
                    "unidade_lotacao": get_value(15),
                    "secretaria_lotacao": get_value(16),
                    "uf": get_value(17),
                    "uorg_exercicio": get_value(18),
                    "unidade_exercicio": get_value(19),
                    "coordenacao": get_value(20),
                    "diretoria": get_value(21),
                    "secretaria": get_value(22),
                    "orgao_origem": get_value(23),
                    "email_institucional": get_value(24),
                    "siape_titular_chefe": get_value(25),
                    "cpf_titular_chefe": get_value(26),
                    "siape_substituto": get_value(27),
                    "cpf_substituto": get_value(28),
                }

                # Processar data de nascimento se disponível
                if len(row) > 3 and not pd.isna(row.iloc[3]):
                    try:
                        # Tentar converter diferentes formatos de data
                        data_nasc = pd.to_datetime(row.iloc[3])
                        dados["data_nascimento"] = data_nasc.date()
                    except:
                        pass  # Se não conseguir converter, deixa em branco

                RelatorioGratificacoes.objects.create(**dados)
                registros_criados += 1

            except Exception as e:
                continue  # Pula registros com erro

        # Marcar como processado
        relatorio_obj.processado = True
        relatorio_obj.data_processamento = timezone.now()
        relatorio_obj.save()

        return True, f"Processamento concluído. {registros_criados} registros criados."

    except Exception as e:
        return False, f"Erro ao processar gratificações: {str(e)}"


def processar_orgaos_centrais(relatorio_obj, arquivo_path):
    """
    Processa planilha de órgãos centrais e setoriais.
    """
    try:
        # Ler planilha Excel com múltiplas abas se necessário
        workbook = openpyxl.load_workbook(arquivo_path)

        # Limpar dados existentes
        RelatorioOrgaosCentrais.objects.filter(relatorio=relatorio_obj).delete()

        registros_criados = 0

        # Processar primeira aba
        sheet = workbook.active

        # Encontrar as seções de órgãos centrais e setoriais
        orgaos_centrais_inicio = None
        orgaos_setoriais_inicio = None

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if row and any(
                cell and "ÓRGÃOS CENTRAIS" in str(cell).upper() for cell in row
            ):
                orgaos_centrais_inicio = i
            elif row and any(
                cell and "ÓRGÃOS SETORIAIS" in str(cell).upper() for cell in row
            ):
                orgaos_setoriais_inicio = i

        # Processar órgãos centrais
        if orgaos_centrais_inicio is not None:
            registros_criados += processar_secao_orgaos(
                sheet,
                orgaos_centrais_inicio,
                orgaos_setoriais_inicio,
                relatorio_obj,
                "central",
            )

        # Processar órgãos setoriais
        if orgaos_setoriais_inicio is not None:
            registros_criados += processar_secao_orgaos(
                sheet, orgaos_setoriais_inicio, None, relatorio_obj, "setorial"
            )

        # Marcar como processado
        relatorio_obj.processado = True
        relatorio_obj.data_processamento = timezone.now()
        relatorio_obj.save()

        return True, f"Processamento concluído. {registros_criados} registros criados."

    except Exception as e:
        return False, f"Erro ao processar órgãos: {str(e)}"


def processar_secao_orgaos(sheet, inicio, fim, relatorio_obj, tipo_orgao):
    """
    Processa uma seção específica da planilha de órgãos.
    """
    registros_criados = 0

    # Encontrar dados de valores nas linhas subsequentes
    linhas_dados = []

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i <= inicio:
            continue
        if fim and i >= fim:
            break

        # Procurar por linhas com dados de cargos e valores
        if row and len(row) >= 2:
            nivel_cargo = row[0]
            valor = row[1]

            if (
                nivel_cargo
                and str(nivel_cargo).strip()
                and valor is not None
                and str(valor).replace(".", "").replace(",", "").isdigit()
            ):

                try:
                    # Converter valor para decimal
                    valor_decimal = float(str(valor).replace(",", "."))

                    RelatorioOrgaosCentrais.objects.create(
                        relatorio=relatorio_obj,
                        tipo_orgao=tipo_orgao,
                        nivel_cargo=str(nivel_cargo).strip(),
                        valor_maximo=valor_decimal,
                        efeitos_financeiros_data="1º DE MAIO DE 2023",  # Padrão baseado na imagem
                    )
                    registros_criados += 1

                except Exception as e:
                    continue

    return registros_criados


def processar_efetivo(relatorio_obj, arquivo_path):
    """
    Processa planilha de efetivo de funcionários.
    """
    try:
        # Ler o arquivo Excel
        df = pd.read_excel(arquivo_path)

        # Limpar dados existentes
        RelatorioEfetivo.objects.filter(relatorio=relatorio_obj).delete()

        registros_criados = 0

        # Processar dados baseado na estrutura vista nas imagens
        for _, row in df.iterrows():
            try:
                # Pular linhas vazias ou de cabeçalho
                if pd.isna(row.iloc[0]) or not str(row.iloc[0]).isdigit():
                    continue

                dados = {
                    "relatorio": relatorio_obj,
                    "qt": int(row.iloc[0]),
                    "nome_completo": (
                        str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ""
                    ),
                    "funcao": str(row.iloc[2]) if not pd.isna(row.iloc[2]) else "",
                    "unidade_macro": (
                        str(row.iloc[3])
                        if len(row) > 3 and not pd.isna(row.iloc[3])
                        else ""
                    ),
                    "horario": (
                        str(row.iloc[4])
                        if len(row) > 4 and not pd.isna(row.iloc[4])
                        else ""
                    ),
                    "bloco_andar": (
                        str(row.iloc[5])
                        if len(row) > 5 and not pd.isna(row.iloc[5])
                        else ""
                    ),
                }

                RelatorioEfetivo.objects.create(**dados)
                registros_criados += 1

            except Exception as e:
                continue

        # Marcar como processado
        relatorio_obj.processado = True
        relatorio_obj.data_processamento = timezone.now()
        relatorio_obj.save()

        return True, f"Processamento concluído. {registros_criados} registros criados."

    except Exception as e:
        return False, f"Erro ao processar efetivo: {str(e)}"


def obter_estatisticas_relatorio(relatorio_obj):
    """
    Retorna estatísticas sobre um relatório processado.
    """
    if not relatorio_obj.processado:
        return None

    stats = {}

    if relatorio_obj.tipo == "gratificacoes":
        stats["total_funcionarios"] = relatorio_obj.dados_gratificacoes.count()
        stats["cargos_unicos"] = (
            relatorio_obj.dados_gratificacoes.values("cargo").distinct().count()
        )
        stats["unidades_unicas"] = (
            relatorio_obj.dados_gratificacoes.values("unidade_lotacao")
            .distinct()
            .count()
        )

    elif relatorio_obj.tipo == "orgaos":
        stats["total_orgaos"] = relatorio_obj.dados_orgaos.count()
        stats["orgaos_centrais"] = relatorio_obj.dados_orgaos.filter(
            tipo_orgao="central"
        ).count()
        stats["orgaos_setoriais"] = relatorio_obj.dados_orgaos.filter(
            tipo_orgao="setorial"
        ).count()

    elif relatorio_obj.tipo in ["efetivo", "facilities"]:
        stats["total_funcionarios"] = relatorio_obj.dados_efetivo.count()
        stats["funcoes_unicas"] = (
            relatorio_obj.dados_efetivo.values("funcao").distinct().count()
        )
        stats["unidades_unicas"] = (
            relatorio_obj.dados_efetivo.values("unidade_macro").distinct().count()
        )

    return stats
