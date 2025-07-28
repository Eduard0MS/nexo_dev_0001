import pandas as pd
from .models import UnidadeCargo, CargoSIORG
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models import PlanilhaImportada
from collections import defaultdict
from openpyxl.styles import Alignment


def processa_planilhas(file_hierarquia, file_estrutura_viva):
    # Leitura da planilha de hierarquia
    if file_hierarquia.name.endswith(".csv"):
        df_hierarquia = pd.read_csv(file_hierarquia, encoding="utf-8")
    else:
        df_hierarquia = pd.read_excel(file_hierarquia)

    # Leitura da planilha de estrutura viva
    if file_estrutura_viva.name.endswith(".csv"):
        df_estrutura_viva = pd.read_csv(file_estrutura_viva, encoding="utf-8")
    else:
        df_estrutura_viva = pd.read_excel(file_estrutura_viva)

    # Processar a planilha de hierarquia: remover metadados e renomear colunas
    df_hierarquia = df_hierarquia.iloc[3:].reset_index(drop=True)
    df_hierarquia.columns = ["Código", "Unidade Organizacional - Sigla"]
    df_hierarquia.dropna(inplace=True)
    df_hierarquia["Código"] = df_hierarquia["Código"].astype(str).str.strip()

    # Processar a planilha de estrutura viva sem alterar os demais dados
    if "Código Unidade" not in df_estrutura_viva.columns:
        raise KeyError(
            "A coluna 'Código Unidade' não foi encontrada na planilha de estrutura viva."
        )
    df_estrutura_viva["Código Unidade"] = (
        df_estrutura_viva["Código Unidade"].astype(str).str.strip()
    )
    if "Categoria" in df_estrutura_viva.columns:
        df_estrutura_viva["Categoria"] = (
            df_estrutura_viva["Categoria"].fillna(0).astype(int)
        )
    if "Nível" in df_estrutura_viva.columns:
        df_estrutura_viva["Nível"] = df_estrutura_viva["Nível"].fillna(0).astype(int)

    # -------------------------------
    # Apenas para construir o campo "Grafo"
    # -------------------------------
    hierarquia_info = {}
    stack = []
    for _, row in df_hierarquia.iterrows():
        codigo = row["Código"]
        unidade = row["Unidade Organizacional - Sigla"]
        # Aqui a lógica calcula o nível para construir o grafo a partir da indentação;
        # ela não altera nenhum dado que será carregado no BD.
        nivel_hierarquico = (len(unidade) - len(unidade.lstrip())) // 5
        while len(stack) > nivel_hierarquico:
            stack.pop()
        if stack:
            grafo_val = f"{hierarquia_info[stack[-1]]['grafo']}-{codigo}"
        else:
            grafo_val = codigo
        stack.append(codigo)
        hierarquia_info[codigo] = {
            "grafo": grafo_val,
            "nivel_hierarquico": nivel_hierarquico,
            "deno_unidade": unidade.strip(),
        }

    # Cria uma cópia dos dados originais da planilha de estrutura viva
    df_resultado = df_estrutura_viva.copy()

    # Atualiza (ou adiciona) as colunas necessárias
    df_resultado["Grafo"] = df_resultado["Código Unidade"].map(
        lambda code: hierarquia_info.get(code, {}).get("grafo", "")
    )
    df_resultado["Nível Hierárquico"] = df_resultado["Código Unidade"].map(
        lambda code: hierarquia_info.get(code, {}).get("nivel_hierarquico", 0)
    )
    df_resultado["Deno Unidade"] = df_resultado["Código Unidade"].map(
        lambda code: hierarquia_info.get(code, {}).get("deno_unidade", "")
    )

    # Garantir que todas as colunas necessárias existam
    colunas_padrao = {
        "Tipo Unidade": "",
        "Sigla Unidade": "",
        "Categoria Unidade": "",
        "Órgão/Entidade": "",
        "Tipo do Cargo": "",
        "Denominação": "",
        "Complemento Denominação": "",
        "Categoria": 0,
        "Nível": 0,
        "Quantidade": 0,
        "Sigla": "",
    }

    for coluna, valor_padrao in colunas_padrao.items():
        if coluna not in df_resultado.columns:
            df_resultado[coluna] = valor_padrao

    # IMPORTANTE: Filtra apenas os registros que possuem um Grafo válido
    # Esses são os registros que realmente fazem parte da estrutura do ministério
    df_resultado = df_resultado[df_resultado["Grafo"].str.strip() != ""]
    print(f"Total de registros após filtragem: {len(df_resultado)}")

    return df_resultado


def salvar_dados_no_banco(df_resultado):
    """
    Salva os dados processados da planilha no banco de dados UnidadeCargo.
    """
    print(f"Iniciando salvamento no banco de dados...")
    print(f"Total de registros a processar: {len(df_resultado)}")
    
    # Limpar dados existentes antes de importar novos
    registros_anteriores = UnidadeCargo.objects.count()
    if registros_anteriores > 0:
        UnidadeCargo.objects.all().delete()
        print(f"Removidos {registros_anteriores} registros anteriores")
    
    registros_criados = 0
    erros = []
    
    for index, row in df_resultado.iterrows():
        try:
            # Garantir que todos os campos obrigatórios tenham valores válidos
            nivel_hierarquico = int(row.get('Nível Hierárquico', 0)) if pd.notna(row.get('Nível Hierárquico')) else 0
            tipo_unidade = str(row.get('Tipo Unidade', '')).strip() if pd.notna(row.get('Tipo Unidade')) else ''
            denominacao_unidade = str(row.get('Deno Unidade', '')).strip() if pd.notna(row.get('Deno Unidade')) else ''
            codigo_unidade = str(row.get('Código Unidade', '')).strip() if pd.notna(row.get('Código Unidade')) else ''
            sigla_unidade = str(row.get('Sigla Unidade', '')).strip() if pd.notna(row.get('Sigla Unidade')) else ''
            categoria_unidade = str(row.get('Categoria Unidade', '')).strip() if pd.notna(row.get('Categoria Unidade')) else ''
            orgao_entidade = str(row.get('Órgão/Entidade', '')).strip() if pd.notna(row.get('Órgão/Entidade')) else ''
            tipo_cargo = str(row.get('Tipo do Cargo', '')).strip() if pd.notna(row.get('Tipo do Cargo')) else ''
            denominacao = str(row.get('Denominação', '')).strip() if pd.notna(row.get('Denominação')) else ''
            complemento_denominacao = str(row.get('Complemento Denominação', '')).strip() if pd.notna(row.get('Complemento Denominação')) else ''
            categoria = int(row.get('Categoria', 0)) if pd.notna(row.get('Categoria')) else 0
            nivel = int(row.get('Nível', 0)) if pd.notna(row.get('Nível')) else 0
            quantidade = int(row.get('Quantidade', 0)) if pd.notna(row.get('Quantidade')) else 0
            grafo = str(row.get('Grafo', '')).strip() if pd.notna(row.get('Grafo')) else ''
            sigla = str(row.get('Sigla', '')).strip() if pd.notna(row.get('Sigla')) else ''
            
            # Validar campos obrigatórios
            if not codigo_unidade:
                print(f"Registro {index + 1} ignorado - código unidade vazio")
                continue
                
            if not grafo:
                print(f"Registro {index + 1} ignorado - grafo vazio")
                continue
            
            # Criar o objeto UnidadeCargo
            unidade_cargo = UnidadeCargo(
                nivel_hierarquico=nivel_hierarquico,
                tipo_unidade=tipo_unidade,
                denominacao_unidade=denominacao_unidade,
                codigo_unidade=codigo_unidade,
                sigla_unidade=sigla_unidade,
                categoria_unidade=categoria_unidade,
                orgao_entidade=orgao_entidade,
                tipo_cargo=tipo_cargo,
                denominacao=denominacao,
                complemento_denominacao=complemento_denominacao,
                categoria=categoria,
                nivel=nivel,
                quantidade=quantidade,
                grafo=grafo,
                sigla=sigla
            )
            
            # Salvar o objeto
            unidade_cargo.save()
            registros_criados += 1
            
            if registros_criados % 50 == 0:
                print(f"Processados {registros_criados} registros...")
                
        except Exception as e:
            erro_msg = f"Erro na linha {index + 1}: {str(e)}"
            erros.append(erro_msg)
            print(f"ERRO: {erro_msg}")
            print(f"Dados da linha: {dict(row)}")
    
    print(f"Salvamento concluído! {registros_criados} registros criados.")
    if erros:
        print(f"Total de erros: {len(erros)}")
        for erro in erros[:5]:  # Mostrar apenas os primeiros 5 erros
            print(f"  - {erro}")
    
    return registros_criados, erros


def processa_organograma():
    """
    Processa os dados das unidades e cargos em uma estrutura de grafo organizacional.
    Retorna um dicionário com a estrutura hierárquica e informações financeiras.
    """
    # Buscar todos os cargos SIORG para referência de valores
    cargos_siorg = {
        f"{cargo.cargo}": {
            "valor": Decimal(
                cargo.valor.replace("R$ ", "").replace(".", "").replace(",", ".")
            ),
            "unitario": cargo.unitario,
        }
        for cargo in CargoSIORG.objects.all()
    }

    # Buscar todas as unidades - filtrando apenas as que têm grafo válido
    unidades = UnidadeCargo.objects.exclude(grafo__exact="").exclude(grafo__isnull=True)

    print(f"Total de unidades com grafo válido: {unidades.count()}")

    # Estrutura para armazenar o organograma
    organograma = {}

    # Primeiro passo: agrupar unidades por código de grafo
    unidades_por_grafo = {}
    for unidade in unidades:
        if not unidade.grafo or unidade.grafo.strip() == "":
            continue

        codigo_atual = unidade.grafo.split("-")[-1]
        if codigo_atual not in unidades_por_grafo:
            unidades_por_grafo[codigo_atual] = []
        unidades_por_grafo[codigo_atual].append(unidade)

    # Segundo passo: processar cada grupo de unidades
    for codigo_atual, grupo_unidades in unidades_por_grafo.items():
        # Ordenar unidades pelo nível do cargo (decrescente)
        grupo_unidades.sort(key=lambda x: x.nivel, reverse=True)

        # Usar a unidade com o cargo de maior nível
        unidade_principal = grupo_unidades[0]
        niveis = unidade_principal.grafo.split("-")

        # Informações do cargo
        cargo_info = {
            "tipo": unidade_principal.tipo_cargo,
            "categoria": unidade_principal.categoria,
            "nivel": unidade_principal.nivel,
            "quantidade": unidade_principal.quantidade,
        }

        # Calcular valor do cargo
        cargo_key = f"{unidade_principal.tipo_cargo} {unidade_principal.categoria} {unidade_principal.nivel:02d}"
        if cargo_key in cargos_siorg:
            cargo_info["valor"] = cargos_siorg[cargo_key]["valor"]
            cargo_info["pontos"] = cargos_siorg[cargo_key]["unitario"]
        else:
            cargo_info["valor"] = Decimal("0.00")
            cargo_info["pontos"] = Decimal("0.00")

        # Criar ou atualizar entrada no organograma
        if codigo_atual not in organograma:
            organograma[codigo_atual] = {
                "codigo": codigo_atual,
                "denominacao": unidade_principal.denominacao_unidade,
                "nivel_hierarquico": unidade_principal.nivel_hierarquico,
                "cargos": [],
                "subordinados": [],
                "pai": niveis[-2] if len(niveis) > 1 else None,
            }

        # Adicionar todos os cargos do mesmo código
        for unidade in grupo_unidades:
            cargo_info = {
                "tipo": unidade.tipo_cargo,
                "categoria": unidade.categoria,
                "nivel": unidade.nivel,
                "quantidade": unidade.quantidade,
            }

            cargo_key = f"{unidade.tipo_cargo} {unidade.categoria} {unidade.nivel:02d}"
            if cargo_key in cargos_siorg:
                cargo_info["valor"] = cargos_siorg[cargo_key]["valor"]
                cargo_info["pontos"] = cargos_siorg[cargo_key]["unitario"]
            else:
                cargo_info["valor"] = Decimal("0.00")
                cargo_info["pontos"] = Decimal("0.00")

            organograma[codigo_atual]["cargos"].append(cargo_info)

        # Estabelecer relações hierárquicas
        if len(niveis) > 1:
            pai = niveis[-2]
            if (
                pai in organograma
                and codigo_atual not in organograma[pai]["subordinados"]
            ):
                organograma[pai]["subordinados"].append(codigo_atual)

    return organograma


def estrutura_json_organograma():
    """
    Estrutura os dados das unidades e cargos em um formato JSON hierárquico.
    Retorna uma lista de unidades com seus cargos e valores.
    """
    from .models import UnidadeCargo, CargoSIORG
    from decimal import Decimal

    # Buscar todos os cargos SIORG para referência de valores
    cargos_siorg = {
        f"{cargo.cargo}": {
            "valor": Decimal(
                cargo.valor.replace("R$ ", "").replace(".", "").replace(",", ".")
            ),
            "unitario": cargo.unitario,
        }
        for cargo in CargoSIORG.objects.all()
    }

    # Buscar todas as unidades com grafo válido
    unidades = UnidadeCargo.objects.exclude(grafo__exact="").exclude(grafo__isnull=True)

    # Lista para armazenar todas as unidades processadas
    unidades_processadas = []

    # Processar cada unidade
    for unidade in unidades:
        # Calcular valores do cargo
        cargo_key = f"{unidade.tipo_cargo} {unidade.categoria} {unidade.nivel:02d}"
        cargo_siorg = cargos_siorg.get(
            cargo_key, {"valor": Decimal("0"), "unitario": Decimal("0")}
        )

        valor_unitario = cargo_siorg["valor"]
        pontos = cargo_siorg["unitario"]

        # Calcular totais
        quantidade = unidade.quantidade or 0
        gasto_total = valor_unitario * quantidade
        pontos_total = pontos * quantidade

        # Estrutura da unidade
        dados_unidade = {
            "tipo_unidade": unidade.tipo_unidade,
            "denominacao_unidade": unidade.denominacao_unidade,
            "codigo_unidade": unidade.codigo_unidade,
            "sigla": unidade.sigla_unidade,
            "tipo_cargo": unidade.tipo_cargo,
            "denominacao": unidade.denominacao,
            "categoria": unidade.categoria,
            "nivel": unidade.nivel,
            "quantidade": quantidade,
            "grafo": unidade.grafo,
            "valor_unitario": float(valor_unitario),
            "pontos": float(pontos),
            "gasto_total": float(gasto_total),
            "pontos_total": float(pontos_total),
        }

        unidades_processadas.append(dados_unidade)

    return unidades_processadas


def processa_json_organograma(json_data):
    """
    Processa os dados do arquivo JSON do organograma e combina com os dados do SIORG.
    Retorna uma estrutura hierárquica com informações detalhadas de cada unidade.
    """
    from .models import CargoSIORG
    from decimal import Decimal
    import json

    # Carregar dados do SIORG para referência
    cargos_siorg = {
        f"{cargo.cargo}": {
            "valor": Decimal(
                cargo.valor.replace("R$ ", "").replace(".", "").replace(",", ".")
            ),
            "unitario": cargo.unitario,
        }
        for cargo in CargoSIORG.objects.all()
    }

    def calcula_valores_unidade(unidade_data):
        """Calcula os valores totais de uma unidade"""
        cargo_key = f"{unidade_data.get('tipo_cargo', '')} {unidade_data.get('categoria', '')} {unidade_data.get('nivel', ''):02d}"
        cargo_info = cargos_siorg.get(
            cargo_key, {"valor": Decimal("0"), "unitario": Decimal("0")}
        )

        quantidade = unidade_data.get("quantidade", 0)
        valor_total = cargo_info["valor"] * quantidade
        pontos_total = cargo_info["unitario"] * quantidade

        return {
            "valor_total": float(valor_total),
            "pontos_total": float(pontos_total),
            "valor_unitario": float(cargo_info["valor"]),
            "pontos_unitario": float(cargo_info["unitario"]),
        }

    def processa_unidade(unidade_data):
        """Processa uma unidade e seus subordinados recursivamente"""
        valores = calcula_valores_unidade(unidade_data)

        dados_unidade = {
            "id": unidade_data.get("id"),
            "codigo": unidade_data.get("codigo_unidade"),
            "nome": unidade_data.get("denominacao_unidade"),
            "sigla": unidade_data.get("sigla_unidade", ""),
            "cargo": {
                "tipo": unidade_data.get("tipo_cargo", ""),
                "categoria": unidade_data.get("categoria", 0),
                "nivel": unidade_data.get("nivel", 0),
                "quantidade": unidade_data.get("quantidade", 0),
                "valor_unitario": valores["valor_unitario"],
                "pontos_unitario": valores["pontos_unitario"],
            },
            "valores": {
                "valor_total": valores["valor_total"],
                "pontos_total": valores["pontos_total"],
            },
            "children": [],
        }

        # Processar unidades subordinadas
        for subordinado in unidade_data.get("subordinados", []):
            sub_dados = processa_unidade(subordinado)
            dados_unidade["children"].append(sub_dados)
            # Somar valores dos subordinados
            dados_unidade["valores"]["valor_total"] += sub_dados["valores"][
                "valor_total"
            ]
            dados_unidade["valores"]["pontos_total"] += sub_dados["valores"][
                "pontos_total"
            ]

        return dados_unidade

    # Processar todo o organograma
    try:
        if isinstance(json_data, str):
            dados = json.loads(json_data)
        else:
            dados = json_data

        return processa_unidade(dados)
    except Exception as e:
        print(f"Erro ao processar JSON do organograma: {str(e)}")
        return None


def _group_identical_cargos(items):
    """
    Agrupa cargos que têm a mesma denominacao e fce/cce (tipo_cargo + categoria + nivel),
    somando suas quantidades.
    """
    grouped = {}
    
    for item in items:
        # Chave única para agrupamento: denominacao + fce/cce
        denominacao = item.get('denominacao', '')
        tipo_cargo = item.get('tipo_cargo', '')
        categoria = item.get('categoria', '')
        nivel = item.get('nivel', '')
        
        # Criar chave única para agrupamento
        key = f"{denominacao}|{tipo_cargo}|{categoria}|{nivel}"
        
        if key in grouped:
            # Somar quantidade ao grupo existente
            grouped[key]['quantidade'] += item.get('quantidade', 0)
        else:
            # Criar novo grupo
            grouped[key] = item.copy()
    
    return list(grouped.values())

def _remove_sigla_from_denominacao(denominacao):
    """
    Remove a sigla do final do nome da unidade (ex: " - SAGE").
    """
    if not denominacao:
        return denominacao
    
    # Remove sigla no formato " - SIGLA" ou " - SIGLA_UNIDADE"
    if ' - ' in denominacao:
        return denominacao.split(' - ')[0].strip()
    
    return denominacao

def _prepare_data_for_excel(data_list):
    """
    Prepara os dados para exportação em Excel com formatação hierárquica.
    """
    processed_list = []
    if not data_list:
        return processed_list
    
    # Debug: log first few items
    print(f"DEBUG: _prepare_data_for_excel received {len(data_list)} items")
    if data_list:
        print(f"DEBUG: First item keys: {list(data_list[0].keys())}")
        print(f"DEBUG: First item: {data_list[0]}")
    
    # Check if this is complete data (327+ items) or filtered data
    is_complete_data = len(data_list) > 100  # Threshold to determine if it's complete data
    
    if is_complete_data:
        print(f"DEBUG: Processing COMPLETE DATA ({len(data_list)} items)")
        return _prepare_complete_data_for_excel(data_list)
    else:
        print(f"DEBUG: Processing FILTERED DATA ({len(data_list)} items)")
        return _prepare_filtered_data_for_excel(data_list)


def _prepare_filtered_data_for_excel(data_list):
    """
    Prepare filtered data for Excel export with hierarchical structure
    """
    print(f"DEBUG: Processing FILTERED DATA ({len(data_list)} items)")
    
    if not data_list:
        return []
    
    # Debug first item
    first_item = data_list[0]
    print(f"DEBUG: First item keys: {list(first_item.keys())}")
    print(f"DEBUG: First item: {first_item}")
    
    # SIMPLIFIED LOGIC: Process ALL items directly
    # Group by denominacao_unidade to organize by units
    units = {}
    for item in data_list:
        denominacao_unidade = item.get('denominacao_unidade', '')
        if denominacao_unidade not in units:
            units[denominacao_unidade] = []
        units[denominacao_unidade].append(item)
    
    print(f"DEBUG: Found {len(units)} units")
    
    processed_list = []
    
    # Sort units to ensure parent units (SUBSECRETARIA) come first
    sorted_units = sorted(units.items(), key=lambda x: (
        # Parent units (containing SUBSECRETARIA) come first (0), others come after (1)
        0 if 'SUBSECRETARIA' in _remove_sigla_from_denominacao(x[0]).upper() else 1,
        # Then alphabetically
        _remove_sigla_from_denominacao(x[0]).lower()
    ))
    
    # Process each unit
    for denominacao_unidade, unit_items in sorted_units:
        # Group identical cargos within this unit
        unit_items = _group_identical_cargos(unit_items)
        
        # Sort by level (higher first), then alphabetically
        unit_items.sort(key=lambda x: (
            -(x.get('nivel', 0) or 0),  # Higher level first (negative for descending)
            x.get('denominacao', '').lower()  # Then alphabetically
        ))
        
        if unit_items:
            # Determine if this is a parent unit or subunit
            first_item = unit_items[0]
            denominacao_unidade_clean = _remove_sigla_from_denominacao(denominacao_unidade)
            
            # Check if this is a parent unit (contains SUBSECRETARIA)
            if 'SUBSECRETARIA' in denominacao_unidade_clean.upper():
                # Parent unit - use full name in UPPERCASE
                unit_name = denominacao_unidade_clean.upper()
            else:
                # Subunit - use first word in lowercase
                unit_name = denominacao_unidade_clean.split()[0].lower()
            
            # Add first job on the same line as the unit
            first_job = unit_items[0]
            tipo_cargo = first_job.get('tipo_cargo', '')
            categoria = first_job.get('categoria', '')
            nivel = first_job.get('nivel', '')
            
            if tipo_cargo and categoria and nivel:
                cargo_formatado = f"{tipo_cargo} {categoria}.{str(nivel).zfill(2)}"
            else:
                cargo_formatado = tipo_cargo or ''
            
            processed_list.append({
                'area': unit_name,
                'quantidade': first_job.get('quantidade', 1),
                'denominacao': first_job.get('denominacao', ''),
                'cargo_formatado': cargo_formatado
            })
            
            # Add remaining jobs on separate lines
            for cargo in unit_items[1:]:
                tipo_cargo = cargo.get('tipo_cargo', '')
                categoria = cargo.get('categoria', '')
                nivel = cargo.get('nivel', '')
                
                if tipo_cargo and categoria and nivel:
                    cargo_formatado = f"{tipo_cargo} {categoria}.{str(nivel).zfill(2)}"
                else:
                    cargo_formatado = tipo_cargo or ''
                
                processed_list.append({
                    'area': '',  # Empty for jobs under unit
                    'quantidade': cargo.get('quantidade', 1),
                    'denominacao': cargo.get('denominacao', ''),
                    'cargo_formatado': cargo_formatado
                })
    
    # Add blank line at the end
    processed_list.append({
        'area': '',
        'quantidade': '',
        'denominacao': '',
        'cargo_formatado': ''
    })
    
    print(f"DEBUG: Processed {len(processed_list)} items")
    return processed_list


def _prepare_complete_data_for_excel(data_list):
    """
    Prepares complete data (327+ items) for Excel using simplified hierarchical logic.
    For complete data, we organize by major organizational units first.
    """
    # For now, use the same logic as filtered data
    return _prepare_filtered_data_for_excel(data_list)


def gerar_anexo_simulacao(data_atual, data_nova):
    """
    Generates an Excel anexo with simulation data.
    Clears rows 5-327 in 'ComparativoEstruturas' sheet,
    and populates data from row 8.
    """
    import os
    
    planilha_ativa = None
    
    # Try to get active planilha
    try:
        planilha_ativa = PlanilhaImportada.objects.get(ativo=True)
        # Verify if the file actually exists
        if not os.path.exists(planilha_ativa.arquivo.path):
            print(f"WARNING: Active planilha file not found: {planilha_ativa.arquivo.path}")
            planilha_ativa = None
    except PlanilhaImportada.DoesNotExist:
        print("WARNING: No active planilha found in database.")
        planilha_ativa = None
    except PlanilhaImportada.MultipleObjectsReturned:
        print("WARNING: Multiple active planilhas found. Using the first one.")
        try:
            planilha_ativa = PlanilhaImportada.objects.filter(ativo=True).first()
        except:
            planilha_ativa = None

    # If no active planilha or file doesn't exist, try to find any available planilha
    if not planilha_ativa:
        print("INFO: Searching for any available planilha template...")
        available_planilhas = PlanilhaImportada.objects.all().order_by('-data_importacao')
        
        for planilha in available_planilhas:
            if os.path.exists(planilha.arquivo.path):
                print(f"INFO: Using available planilha: {planilha.nome}")
                planilha_ativa = planilha
                # Set this one as active for future use
                planilha.ativo = True
                planilha.save()
                break
        
        if not planilha_ativa:
            raise FileNotFoundError(
                "Nenhum template de planilha encontrado no sistema. "
                "Por favor, importe um arquivo de template através do admin."
            )

    try:
        # Load the workbook, trying with keep_vba=False first
        workbook = openpyxl.load_workbook(
            planilha_ativa.arquivo.path, keep_vba=False, data_only=False
        )
        print(f"SUCCESS: Loaded template: {planilha_ativa.nome}")
    except Exception as e:
        raise ValueError(f"Erro ao carregar o template da planilha '{planilha_ativa.nome}': {str(e)}")

    sheet_name = "ComparativoEstruturas"
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada no template.")

    sheet = workbook[sheet_name]

    # 1. Clear content of rows 5 to 327, handling merged cells
    for row_index in range(5, 328):
        if row_index <= sheet.max_row:
            for col_index in range(1, sheet.max_column + 1):
                if col_index <= sheet.max_column:
                    cell = sheet.cell(row=row_index, column=col_index)

                    # Check if the cell is part of a merged cell range
                    is_merged = False
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            range_to_unmerge_str = str(merged_range)
                            sheet.unmerge_cells(range_to_unmerge_str)
                            break

                    try:
                        cell.value = None
                    except AttributeError:
                        pass

    # 2. Headers are already in the template on row 7 - DO NOT ADD THEM AGAIN
    #    Data will be populated starting from row 8.

    # Debug: log input data
    print(f"DEBUG: gerar_anexo_simulacao - data_atual length: {len(data_atual)}")
    print(f"DEBUG: gerar_anexo_simulacao - data_nova length: {len(data_nova)}")
    
    # Debug: log input data
    print(f"DEBUG: gerar_anexo_simulacao - data_atual length: {len(data_atual)}")
    print(f"DEBUG: gerar_anexo_simulacao - data_nova length: {len(data_nova)}")

    # 3. Prepare and populate data
    processed_atual = _prepare_data_for_excel(data_atual)
    processed_nova = _prepare_data_for_excel(data_nova)
    
    # Debug: log processed data
    print(f"DEBUG: gerar_anexo_simulacao - processed_atual length: {len(processed_atual)}")
    print(f"DEBUG: gerar_anexo_simulacao - processed_nova length: {len(processed_nova)}")
    
    # Debug: log processed data
    print(f"DEBUG: gerar_anexo_simulacao - processed_atual length: {len(processed_atual)}")
    print(f"DEBUG: gerar_anexo_simulacao - processed_nova length: {len(processed_nova)}")

    # Define alignments
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    current_data_row = 8  # Data starts at row 8

    # Populate "Estrutura Atual" (Columns A-D, i.e., 1-4)
    for item_row_data in processed_atual:
        cell_A = sheet.cell(row=current_data_row, column=1)
        cell_A.value = item_row_data["area"]
        cell_A.alignment = align_left

        cell_B = sheet.cell(row=current_data_row, column=2)
        cell_B.value = item_row_data["quantidade"]
        cell_B.alignment = align_center

        cell_C = sheet.cell(row=current_data_row, column=3)
        cell_C.value = item_row_data["denominacao"]
        cell_C.alignment = align_left

        cell_D = sheet.cell(row=current_data_row, column=4)
        cell_D.value = item_row_data["cargo_formatado"]
        cell_D.alignment = align_left

        current_data_row += 1

    current_data_row = 8  # Reset for "Estrutura Nova"

    # Populate "Estrutura Nova" (Columns F-I, i.e., 6-9)
    for item_row_data in processed_nova:
        cell_F = sheet.cell(row=current_data_row, column=6)
        cell_F.value = item_row_data["area"]
        cell_F.alignment = align_left

        cell_G = sheet.cell(row=current_data_row, column=7)
        cell_G.value = item_row_data["quantidade"]
        cell_G.alignment = align_center

        cell_H = sheet.cell(row=current_data_row, column=8)
        cell_H.value = item_row_data["denominacao"]
        cell_H.alignment = align_left

        cell_I = sheet.cell(row=current_data_row, column=9)
        cell_I.value = item_row_data["cargo_formatado"]
        cell_I.alignment = align_left

        current_data_row += 1

    # Save to a BytesIO stream
    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)  # Rewind the stream to the beginning

    return excel_stream


# Make sure to add openpyxl to requirements.txt
# Example: openpyxl>=3.0.0
